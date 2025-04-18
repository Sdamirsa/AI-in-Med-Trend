###########################
### Importing libraries ###
###########################
# %pip install langchain langchain_core langchain_openai langchain_community pandas openpyxl python-docx python-dotenv PyPDF2 nest_asyncio

# system
import asyncio
from asyncio import CancelledError
import os
import time
import platform
import subprocess
from typing import List, Dict, Any, Literal, Union, Optional
import getpass

# Langchain and LLM
import re
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from dotenv import load_dotenv
from pydantic import BaseModel, Field

from langchain_core.prompts import PromptTemplate
from langchain_openai import ChatOpenAI
from langchain_openai import AzureChatOpenAI
from langchain_community.callbacks.manager import get_openai_callback
from langchain.output_parsers import PydanticOutputParser, OutputFixingParser
from langchain_core.messages import AIMessage
from pathlib import Path


##############################
###    Helper Functions   ####
##############################

def ensure_azure_env_vars():
    """
    Ensures that the environment variables for Azure OpenAI 
    are set. If not, prompts the user for values and saves 
    them to the local .env file.
    """
    azure_vars = {
        "AZURE_OPENAI_API_KEY": "Please provide the Azure OpenAI API Key: ",
        "AZURE_OPENAI_ENDPOINT": "Please provide the Azure OpenAI Endpoint: ",
        "AZURE_API_VERSION": "Please provide the Azure OpenAI API Version: "
    }
    updated = False
    
    for var, prompt_msg in azure_vars.items():
        if not os.getenv(var):
            val = input(prompt_msg)
            os.environ[var] = val
            with open(".env", "a", encoding="utf-8") as f:
                f.write(f"{var}={val}\n")
            updated = True

    # Reload .env if new variables were added
    if updated:
        load_dotenv()


def check_environment_keys(env_keys: List[str]) -> None:
    """
    Check if the required keys are provided 
    (non-empty strings after we've ensured environment variables).
    """
    missing_keys = [key for key in env_keys if not key]
    if missing_keys:
        raise ValueError(f"Missing API keys or endpoints: {', '.join(missing_keys)}")


def extract_json(message: AIMessage) -> Any:
    """Extracts JSON content from a string where JSON is 
       embedded between ```json and ``` tags.
    """
    text = message.content
    print(f"THE ERROR IS CAUSED BY: {text}.")
    pattern = r"```json(.*?)```"
    matches = re.findall(pattern, text, re.DOTALL)
    if not matches:
        raise ValueError("No JSON block found in the message.")

    try:
        return json.loads(matches[0].strip())
    except json.JSONDecodeError as e:
        raise ValueError(f"Failed to parse JSON: {e}") from e


#######################################
######      OpenAI + Pydantic    ######
#######################################
async def async_openai_pydantic_extractor(
    input_text: str,
    PydanticClass: BaseModel,
    model: str = 'gpt-3.5-turbo',
    pre_prompt: str = "",
    temperature: float = 0,
    max_tokens: int = 2048,
    max_retries: int = 2,
    parser_error_handling: Literal["include_raw", "llm_to_correct", "manual_logic"]= "llm_to_correct", 
    openai_api_key=os.getenv("SIAVOSH_OPENAI_API_KEY"),
) -> tuple:
    """
    Asynchronous function for extracting information from 
    a text using OpenAI and Pydantic.
    """
    # Make sure the environment variable for OpenAI is set
    check_environment_keys([openai_api_key])
    
    llm = ChatOpenAI(
        api_key=openai_api_key,
        model=model,
        temperature=temperature,
        max_retries=max_retries,
        max_tokens=max_tokens,
    )

    if parser_error_handling == "include_raw":
        parser = PydanticOutputParser(pydantic_object=PydanticClass, include_raw=True)
    elif parser_error_handling == "llm_to_correct":
        first_parser = PydanticOutputParser(pydantic_object=PydanticClass)
        parser = OutputFixingParser.from_llm(parser=first_parser, llm=llm)
    elif parser_error_handling == "manual_logic":
        parser = extract_json
    else:
        raise ValueError(f"Invalid parser_error_handling value: '{parser_error_handling}'.")

    prompt = PromptTemplate(
        template="{pre_prompt}\nAnswer the user query and return JSON.\n{format_instructions}\n{query}".strip(),
        input_variables=["pre_prompt", "query"],
        partial_variables={"format_instructions": parser.get_format_instructions()},
    )

    chain = prompt | llm | parser
    experiment_info = {
        "model": model,
        "pre_prompt": pre_prompt,
        "all_prompt": prompt,
        "temperature": temperature,
        "max_tokens": max_tokens,
        "max_retries": max_retries,
    }

    try:
        with get_openai_callback() as cb:
            start_time = time.time()
            response = await chain.ainvoke({"pre_prompt": pre_prompt, "query": input_text})
            end_time = time.time()
            experiment_info.update({
                "execution_time": round((end_time - start_time), 2),
                "total_tokens": cb.total_tokens,
                "prompt_tokens": cb.prompt_tokens,
                "completion_tokens": cb.completion_tokens,
                "total_cost": cb.total_cost
            })
            status = "EXTRACTED"
    except (ConnectionError, TimeoutError) as e:
        status, response = f"NETWORK ERROR: {e}", None
    except Exception as e:
        status, response = f"ERROR in extraction: {e}", None

    print(f"STATUS of {PydanticClass.__name__}: {status}")
    return PydanticClass.__name__, status, response, experiment_info


#######################################
######   AzureOpenAI + Pydantic   #####
#######################################
async def async_azureOpenAI_pydantic_extractor(
    input_text: str,
    PydanticClass: BaseModel,
    model: str = 'gpt-4o',
    pre_prompt: str = "",
    temperature: float = 0,
    max_tokens: int = 2048,
    logprobs=False,
    seed=None,
    error_handling_strategy: Literal["include_raw", "llm_to_correct", "manual_logic"]= "llm_to_correct",
    timeout=60,
    max_retries: int = 2,
    parser_error_handling: Literal["include_raw", "llm_to_correct", "manual_logic"]= "llm_to_correct",
    azure_api_key=os.getenv("AZURE_OPENAI_API_KEY"),
    azure_endpoint=os.getenv("AZURE_OPENAI_ENDPOINT"),
    azure_api_version=os.getenv("AZURE_API_VERSION"),
    **kwargs,
) -> tuple:
    """
    Asynchronous function for extracting information using Azure OpenAI and Pydantic.
    """
    # Ensure environment variables for Azure are set or prompt user for them
    ensure_azure_env_vars()
    azure_api_key = os.getenv("AZURE_OPENAI_API_KEY", azure_api_key)
    azure_endpoint = os.getenv("AZURE_OPENAI_ENDPOINT", azure_endpoint)
    azure_api_version = os.getenv("AZURE_API_VERSION", azure_api_version)

    check_environment_keys([azure_api_key, azure_endpoint, azure_api_version])

    llm = AzureChatOpenAI(
        api_key=azure_api_key,
        azure_endpoint=azure_endpoint,
        api_version=azure_api_version,
        model=model,
        max_tokens=max_tokens,
        timeout=timeout,
        temperature=temperature,
        logprobs=logprobs,
        seed=seed,
        max_retries=max_retries,
        **kwargs,
    )

    if parser_error_handling == "include_raw":
        parser = PydanticOutputParser(pydantic_object=PydanticClass, include_raw=True)
    elif parser_error_handling == "llm_to_correct":
        first_parser = PydanticOutputParser(pydantic_object=PydanticClass)
        parser = OutputFixingParser.from_llm(parser=first_parser, llm=llm)
    elif parser_error_handling == "manual_logic":
        parser = extract_json
    else:
        raise ValueError(f"Invalid parser_error_handling value: '{parser_error_handling}'.")

    prompt = PromptTemplate(
        template="{pre_prompt}\nAnswer the user query and return JSON.\n{format_instructions}\n{query}".strip(),
        input_variables=["pre_prompt", "query"],
        partial_variables={"format_instructions": parser.get_format_instructions()},
    )

    chain = prompt | llm | parser
    experiment_info = {
        "model": model,
        "pre_prompt": pre_prompt,
        "all_prompt": prompt,
        "temperature": temperature,
        "max_tokens": max_tokens,
        "max_retries": max_retries,
    }

    try:
        with get_openai_callback() as cb:
            start_time = time.time()
            response = await chain.ainvoke({"pre_prompt": pre_prompt, "query": input_text})
            end_time = time.time()
            experiment_info.update({
                "execution_time": round((end_time - start_time), 2),
                "total_tokens": cb.total_tokens,
                "prompt_tokens": cb.prompt_tokens,
                "completion_tokens": cb.completion_tokens,
                "total_cost": cb.total_cost
            })
            status = "EXTRACTED"
    except (ConnectionError, TimeoutError) as e:
        status, response = f"NETWORK ERROR: {e}", None
    except Exception as e:
        status, response = f"ERROR in extraction: {e}", None

    print(f"STATUS of {PydanticClass.__name__}: {status}")
    return PydanticClass.__name__, status, response, experiment_info


#######################################
######     Runpod + Pydantic      #####
#######################################
async def async_runpod_pydantic_extractor(
    input_text: str,
    PydanticClass: BaseModel,
    runpod_base_url:str,
    runpod_api:str,
    model: str = '',
    pre_prompt: str = "",
    temperature: float = 0,
    max_tokens: int = 2048,
    max_retries: int = 2,
    parser_error_handling: Literal["include_raw", "llm_to_correct", "manual_logic"]= "llm_to_correct",
) -> tuple:
    """
    Asynchronous function for extracting information using a 
    Runpod-based LLM and Pydantic.
    """
    check_environment_keys([runpod_api, runpod_base_url])

    llm = ChatOpenAI(
        api_key=runpod_api,
        model=model,
        base_url=runpod_base_url,
        temperature=temperature,
        max_retries=max_retries,
        max_tokens=max_tokens,
    )

    if parser_error_handling == "include_raw":
        parser = PydanticOutputParser(pydantic_object=PydanticClass, include_raw=True)
    elif parser_error_handling == "llm_to_correct":
        first_parser = PydanticOutputParser(pydantic_object=PydanticClass)
        parser = OutputFixingParser.from_llm(parser=first_parser, llm=llm)
    elif parser_error_handling == "manual_logic":
        parser = extract_json
    else:
        raise ValueError(f"Invalid parser_error_handling value: '{parser_error_handling}'.")

    prompt = PromptTemplate(
        template="{pre_prompt}\nAnswer the user query and return JSON.\n{format_instructions}\n{query}".strip(),
        input_variables=["pre_prompt", "query"],
        partial_variables={"format_instructions": parser.get_format_instructions()},
    )

    chain = prompt | llm | parser
    experiment_info = {
        "model": model,
        "pre_prompt": pre_prompt,
        "all_prompt": prompt,
        "temperature": temperature,
        "max_tokens": max_tokens,
        "max_retries": max_retries,
    }

    try:
        with get_openai_callback() as cb:
            start_time = time.time()
            response = await chain.ainvoke({"pre_prompt": pre_prompt, "query": input_text})
            end_time = time.time()
            experiment_info.update({
                "execution_time": round((end_time - start_time), 2),
                "total_tokens": cb.total_tokens,
                "prompt_tokens": cb.prompt_tokens,
                "completion_tokens": cb.completion_tokens,
                "total_cost": cb.total_cost
            })
            status = "EXTRACTED"
    except (ConnectionError, TimeoutError) as e:
        status, response = f"NETWORK ERROR: {e}", None
    except Exception as e:
        status, response = f"ERROR in extraction: {e}", None

    print(f"STATUS of {PydanticClass.__name__}: {status}")
    return PydanticClass.__name__, status, response, experiment_info


#######################################
######    Multiengine handling   ######
#######################################
async def async_EXCT_MARIA_multiengine(
    index: int,
    row: Any,
    model_engine: Literal["OpenAI_Async", "Runpod_Async","AzureOpenAI_Async"],
    input_text: str,
    Pydantic_Objects_List: List[BaseModel],
    parser_error_handling: Literal["include_raw", "llm_to_correct", "manual_logic"]= "llm_to_correct", 
    model: str = 'gpt-3.5-turbo',
    pre_prompt: str = "",
    temperature: float = 0,
    max_tokens: int = 2048,
    logprobs=False,
    seed=None,
    timeout=60,
    max_retries: int = 2,
    openai_api_key=os.getenv("OPENAI_API_KEY"),
    runpod_base_url:str="",
    runpod_api:str=os.getenv("runpod_EXCT_api"),
    azure_api_key=os.getenv("AZURE_OPENAI_API_KEY"),
    azure_endpoint=os.getenv("AZURE_OPENAI_ENDPOINT"),
    azure_api_version=os.getenv("AZURE_API_VERSION"),
) -> tuple:
    """
    Asynchronous function to run multiple extractions with the selected engine.
    """
    EXCT_output_dic = {}
    
    if model_engine == "OpenAI_Async":
        for PydanticClass in Pydantic_Objects_List:
            cls_name, status, response, experiment_info = await async_openai_pydantic_extractor(
                input_text,
                PydanticClass,
                model=model,
                pre_prompt=pre_prompt,
                temperature=temperature,
                max_tokens=max_tokens,
                max_retries=max_retries,
                parser_error_handling=parser_error_handling,
                openai_api_key=openai_api_key
            )
            EXCT_output_dic[cls_name] = {
                "status": status,
                "response": response,
                "experiment_info": experiment_info
            }

    elif model_engine == "Runpod_Async":
        for PydanticClass in Pydantic_Objects_List:
            cls_name, status, response, experiment_info = await async_runpod_pydantic_extractor(
                runpod_base_url=runpod_base_url,
                runpod_api=runpod_api,
                input_text=input_text,
                PydanticClass=PydanticClass,
                model=model,
                pre_prompt=pre_prompt,
                temperature=temperature,
                max_tokens=max_tokens,
                max_retries=max_retries,
                parser_error_handling=parser_error_handling,
            )
            EXCT_output_dic[cls_name] = {
                "status": status,
                "response": response,
                "experiment_info": experiment_info
            }
    
    elif model_engine == "AzureOpenAI_Async":
        for PydanticClass in Pydantic_Objects_List:
            cls_name, status, response, experiment_info = await async_azureOpenAI_pydantic_extractor(
                input_text=input_text,
                PydanticClass=PydanticClass,
                model=model,
                pre_prompt=pre_prompt,
                temperature=temperature,
                max_tokens=max_tokens,
                max_retries=max_retries,
                logprobs=logprobs,
                seed=seed,
                timeout=timeout,
                parser_error_handling=parser_error_handling,
                azure_api_key=azure_api_key,
                azure_endpoint=azure_endpoint,
                azure_api_version=azure_api_version,
            )
            EXCT_output_dic[cls_name] = {
                "status": status,
                "response": response,
                "experiment_info": experiment_info
            }
           
    return index, row, EXCT_output_dic


#######################################
######        Excel Handler      ######
#######################################

def open_excel_file(file_path: str) -> None:
    """
    Open an Excel file based on the operating system.
    """
    if platform.system() == 'Windows':
        os.startfile(file_path)
    elif platform.system() == 'Darwin':  # macOS
        subprocess.call(('open', file_path))
    elif platform.system() == 'Linux':
        subprocess.call(('xdg-open', file_path))


def add_color_to_cell(sheet, cell: str, color: str) -> None:
    """
    Fill a specific cell in Excel with a specified ARGB color.
    """
    if len(color) == 7:  # Convert color to ARGB if only RGB (e.g. #FF0000 -> FFFF0000)
        color = f"FF{color[1:]}"
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    sheet[cell].fill = fill


def color_cells_in_excel(file_path: str, argb_color: str, experiment_name: str) -> None:
    """
    Color cells in an Excel file based on specific conditions 
    (checking a column named after experiment_name).
    """
    wb = load_workbook(file_path)
    sheet = wb.active
    df = pd.read_excel(file_path)

    # Identify relevant columns by experiment name prefix
    experiment_columns = [col for col in df.columns if col.startswith(f"{experiment_name}_")]

    for index, row in df.iterrows():
        if row.get(experiment_name) == "EXTRACTED":
            for col in experiment_columns:
                cell = sheet.cell(row=index + 2, column=df.columns.get_loc(col) + 1)  # +2 for header offset
                add_color_to_cell(sheet, cell.coordinate, argb_color)

    wb.save(file_path)
    print(f"Cells in columns starting with {experiment_name}_ have been colored with {argb_color}.")


def save_results_to_excel_simple(df, index, experiment_label, EXCT_output_dic):
    """
    Flatten the Pydantic or dict responses into a single row 
    in the DataFrame for each PydanticClass.
    """
    updated_df = df.copy()

    def flatten_pydantic(obj, prefix=''):
        flat_dict = {}
        stack = [(obj, prefix)]
        while stack:
            current_obj, current_prefix = stack.pop()

            if isinstance(current_obj, BaseModel):
                current_obj = current_obj.model_dump()  # Convert to dict

            if isinstance(current_obj, dict):
                for key, value in current_obj.items():
                    full_key = f"{current_prefix}_{key}" if current_prefix else key
                    if isinstance(value, (BaseModel, dict)):
                        stack.append((value, full_key))
                        continue
                    flat_dict[full_key] = value
        return flat_dict

    # Update the DataFrame with extracted results
    for key, result in EXCT_output_dic.items():
        response = result.get("response", None)

        if isinstance(response, BaseModel):
            flattened_response = flatten_pydantic(response)
            for field_name, field_value in flattened_response.items():
                column_name = f"{experiment_label}_Genrator_{key}_{field_name}"
                if column_name not in updated_df.columns:
                    updated_df[column_name] = None
                updated_df.at[index, column_name] = str(field_value)

        elif isinstance(response, dict):
            flattened_response = flatten_pydantic(response)
            for field_name, field_value in flattened_response.items():
                column_name = f"{experiment_label}_Genrator_{key}_{field_name}"
                if column_name not in updated_df.columns:
                    updated_df[column_name] = None
                updated_df.at[index, column_name] = str(field_value)

    updated_df.at[index, f"{experiment_label}_Generator_response"] = str(EXCT_output_dic)
    return updated_df


###########################################
## Main Async Function for Multi-row Run ##
###########################################

async def EXCT_main(
    excel_file_path: str,
    experiment_label: str,
    Pydantic_Objects_List: List[BaseModel],
    model_engine: Literal["OpenAI_Async", "Runpod_Async", "AzureOpenAI_Async"] = "OpenAI_Async",
    parser_error_handling: Literal["include_raw", "llm_to_correct", "manual_logic"] = "llm_to_correct",
    open_at_end: bool = True,
    model: str = 'gpt-3.5-turbo',
    pre_prompt: str = "",
    temperature: float = 0,
    max_tokens: int = 2048,
    logprobs: bool = False,
    seed=None,
    timeout: int = 60,
    max_retries: int = 2,
    total_async_n: int = 10,
    openai_api_key: str = os.getenv("OPENAI_API_KEY"),
    runpod_base_url: str = "",
    runpod_api: str = "",
    azure_api_key: str = os.getenv("AZURE_OPENAI_API_KEY"),
    azure_endpoint: str = os.getenv("AZURE_OPENAI_ENDPOINT"),
    azure_api_version: str = os.getenv("AZURE_API_VERSION"),
    text_column: Optional[str] = "All_Report_Text",
    textPath_column: Optional[str] = None,
) -> None:
    """
    Main function to perform asynchronous extraction across multiple rows of a CSV/Excel file.
    - Reads the file (CSV or Excel) into a DataFrame
    - Creates asynchronous tasks for each row
    - Calls the relevant LLM extraction engine 
    - Saves results in a new file (excel_output_path) + JSON
    """
    # 1) Create a new variable to save the final output so as not to overwrite original
    #    We'll name it "excel_output_path".
    base_name, ext = os.path.splitext(excel_file_path)
    excel_output_path = f"{base_name}_output_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
    json_output_path = f"{base_name}_output_{time.strftime('%Y%m%d_%H%M%S')}.json"

    # 2) Check extension and read input CSV or Excel
    if ext.lower() == ".csv":
        df = pd.read_csv(excel_file_path)
    elif ext.lower() in [".xls", ".xlsx"]:
        df = pd.read_excel(excel_file_path)
    else:
        raise ValueError("Unsupported file format. Please provide a CSV (.csv) or Excel (.xls/.xlsx) file.")

    # Validate columns
    if text_column and textPath_column:
        raise ValueError("You should define either a column name for text (text_column), OR a column path to document (textPath_column), but not both.")

    # Add experiment label column if doesn't exist
    if experiment_label not in df.columns:
        df[experiment_label] = ""

    last_row_index = df.shape[0] - 1
    async_tasks = []
    async_n = 0
    processed_count = 0

    try:
        for index, row in df.iterrows():
            if row[experiment_label] != "EXTRACTED":
                try:
                    if text_column: 
                        input_text = row[text_column]
                    elif textPath_column:
                        # If you had a doc reading function, you could do something like:
                        #   input_text = read_document(row[textPath_column])
                        # For now, let's just read text from that column path:
                        with open(row[textPath_column], 'r', encoding='utf-8') as f:
                            input_text = f.read()
                    else:
                        raise ValueError("No valid column for text input was specified.")

                    task = asyncio.create_task(
                        async_EXCT_MARIA_multiengine(
                            index=index,
                            row=row,
                            model_engine=model_engine,
                            input_text=input_text,
                            Pydantic_Objects_List=Pydantic_Objects_List,
                            parser_error_handling=parser_error_handling,
                            model=model,
                            pre_prompt=pre_prompt,
                            temperature=temperature,
                            max_tokens=max_tokens,
                            logprobs=logprobs,
                            seed=seed,
                            timeout=timeout,
                            max_retries=max_retries,
                            openai_api_key=openai_api_key,
                            runpod_base_url=runpod_base_url,
                            runpod_api=runpod_api,
                            azure_api_key=azure_api_key,
                            azure_endpoint=azure_endpoint,
                            azure_api_version=azure_api_version,
                        )
                    )

                    async_tasks.append(task)
                    async_n += 1

                    # Process tasks in batches
                    if async_n == total_async_n or (index == last_row_index and async_tasks):
                        results = await asyncio.gather(*async_tasks)
                        for res in results:
                            the_index, the_row, EXCT_output_dic = res
                            # Save extraction to DF
                            df = save_results_to_excel_simple(df, the_index, experiment_label, EXCT_output_dic)
                            # Mark status as EXTRACTED
                            df.at[the_index, experiment_label] = "EXTRACTED"

                        async_tasks = []
                        async_n = 0
                        processed_count += total_async_n
                        print(f"Finished: {index} of {last_row_index}")

                        # (Optional) Save a backup every 500 processed
                        if processed_count >= 500:
                            backup_file_path = base_name + f"_backup_{index}.xlsx"
                            df.to_excel(backup_file_path, index=False)
                            print(f"Backup saved to {backup_file_path}")
                            processed_count = 0

                except Exception as e:
                    df.at[index, experiment_label] = f"ERROR: {e}"

    except (KeyboardInterrupt, CancelledError):
        print("Process interrupted by user.")

    finally:
        # 3) Save the final DataFrame to Excel (excel_output_path)
        df.to_excel(excel_output_path, index=False)
        print(f"Final data saved to {excel_output_path}")


        try:
            df.to_json(json_output_path, orient="records", force_ascii=False)
            print(f"Final data also saved to JSON: {json_output_path}")
        except Exception as json_err:
            print(f"Error occurred while saving JSON: {json_err}")

        # Optionally open the file after saving
        if open_at_end:
            open_excel_file(excel_output_path)
            
            
            
            
import asyncio
import json
import os
import time
from typing import List, Dict, Any, Literal, Union, Optional

# 1) Import the relevant pieces from your existing code:
#    (Either place this code in the same file, or import these methods
#     from your S3_EXCT_Function.py)
#
#    - async_EXCT_MARIA_multiengine
#    - The various Pydantic extraction functions
#    - ensure_azure_env_vars, check_environment_keys, etc. if needed

############################################################
###  OPTIONAL: Helper function to traverse a path in JSON ###
############################################################

def get_json_sublist(data: Union[list, dict], path_to_list: List[str]) -> Any:
    """
    Given a JSON object (list or dict) and a list of keys (path_to_list),
    traverse down that path to retrieve a sub-element.
    
    Example:
       data = {"foo": {"bar": [{"name": "Alice"}, {"name": "Bob"}]}}
       path_to_list = ["foo", "bar"]
       => returns the list [{"name": "Alice"}, {"name": "Bob"}]
    """
    current = data
    for key in path_to_list:
        if isinstance(current, dict) and key in current:
            current = current[key]
        else:
            raise ValueError(f"Invalid path segment '{key}' in {path_to_list}.")
    return current

##############################################################
###  Function to assemble an "EXCT" dictionary from results ###
##############################################################

def build_exct_dictionary(EXCT_output_dic: dict) -> dict:
    """
    Takes the return from async_EXCT_MARIA_multiengine (a dict of
    {PydanticClassName: {...}}) and creates a single "EXCT" dictionary.

    The user wants:
      {
        "EXCT": {
          "status": "...",
          "cost": ...,
          "PydanticClassName1": {...fields...},
          "PydanticClassName2": {...fields...},
          ...
        }
      }

    We'll set "status" to "EXTRACTED" if everything succeeded or store
    an error status if something failed. We can also sum up the total cost.
    Adjust as needed.
    """
    exct_data = {}
    total_cost = 0.0
    all_good = True

    for pyd_cls_name, result in EXCT_output_dic.items():
        status = result.get("status", "")
        experiment_info = result.get("experiment_info", {})
        response = result.get("response", None)

        # If there's an error in any pydantic object, we can reflect that
        if not status.startswith("EXTRACTED"):
            all_good = False

        # Collect cost from experiment_info if present
        cost_val = experiment_info.get("total_cost", 0.0)
        if isinstance(cost_val, (int, float)):
            total_cost += cost_val

        # We store the actual fields from the response if it's a dict or pydantic object
        # (If using pydantic, you can do response.model_dump() if needed.)
        if hasattr(response, "model_dump"):
            exct_data[pyd_cls_name] = response.model_dump()
        elif isinstance(response, dict):
            exct_data[pyd_cls_name] = response
        else:
            # fallback: store raw
            exct_data[pyd_cls_name] = {"raw_response": str(response)}

    exct_data["status"] = "EXTRACTED" if all_good else "ERROR"
    exct_data["cost"] = round(total_cost, 5)
    return exct_data

#######################################################################
###  Main Async JSON function that reads JSON, does extraction, and  ###
###  writes back to "EXCT" in each item                              ###
#######################################################################

async def EXCT_main_JSON(
    json_file_path: str,
    text_key: str,
    Pydantic_Objects_List: List[Any],  # your pydantic classes
    path_to_list: Optional[List[str]] = None,
    model_engine: Literal["OpenAI_Async", "Runpod_Async", "AzureOpenAI_Async"] = "OpenAI_Async",
    parser_error_handling: Literal["include_raw", "llm_to_correct", "manual_logic"] = "llm_to_correct",
    model: str = 'gpt-3.5-turbo',
    pre_prompt: str = "",
    temperature: float = 0,
    max_tokens: int = 2048,
    logprobs: bool = False,
    seed=None,
    timeout: int = 60,
    max_retries: int = 2,
    openai_api_key: str = os.getenv("OPENAI_API_KEY"),
    runpod_base_url: str = "",
    runpod_api: str = "",
    azure_api_key: str = os.getenv("AZURE_OPENAI_API_KEY"),
    azure_endpoint: str = os.getenv("AZURE_OPENAI_ENDPOINT"),
    azure_api_version: str = os.getenv("AZURE_API_VERSION"),
    total_async_n: int = 5,
    output_file_path: Optional[str] = None,
) -> None:
    """
    1) Loads JSON from 'json_file_path'.
    2) If path_to_list is provided, we traverse the JSON to that location (which should be a list).
       If no path is provided, we assume the loaded JSON is itself a list.
    3) For each item in the list, call the LLM extraction function (async_EXCT_MARIA_multiengine).
       The text fed to the model is item[text_key].
    4) The results are collected and stored in item["EXCT"] = { ... }.
    5) Write the updated JSON to disk (or to 'output_file_path' if given).
    """

    # --- 1) Read the JSON ---
    with open(json_file_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # --- 2) Locate the list we want to loop over ---
    if path_to_list:
        items = get_json_sublist(data, path_to_list)
        if not isinstance(items, list):
            raise ValueError("The path_to_list does not point to a list in the JSON.")
    else:
        if isinstance(data, list):
            items = data
        else:
            raise ValueError(
                "No path_to_list was provided, and top-level JSON is not a list."
            )

    # Prepare asynchronous tasks
    async_tasks = []
    async_n = 0
    last_index = len(items) - 1

    # We'll track results so we can place them back in the item
    # once tasks complete.
    for index, item in enumerate(items):
        # If we already have item["EXCT"], skip or re-run? 
        # For now, we do not skip—modify as you wish.
        text_input = item.get(text_key, None)
        if not text_input:
            # If there's no text, we can skip or store an error
            item["EXCT"] = {"status": f"No '{text_key}' found in item."}
            continue

        task = asyncio.create_task(
            async_EXCT_MARIA_multiengine(
                index=index,
                row=None,  # Not needed for JSON
                model_engine=model_engine,
                input_text=text_input,
                Pydantic_Objects_List=Pydantic_Objects_List,
                parser_error_handling=parser_error_handling,
                model=model,
                pre_prompt=pre_prompt,
                temperature=temperature,
                max_tokens=max_tokens,
                logprobs=logprobs,
                seed=seed,
                timeout=timeout,
                max_retries=max_retries,
                openai_api_key=openai_api_key,
                runpod_base_url=runpod_base_url,
                runpod_api=runpod_api,
                azure_api_key=azure_api_key,
                azure_endpoint=azure_endpoint,
                azure_api_version=azure_api_version,
            )
        )
        async_tasks.append(task)
        async_n += 1

        # We process in batches of total_async_n
        if async_n == total_async_n or (index == last_index and async_tasks):
            results = await asyncio.gather(*async_tasks)
            for res in results:
                the_index, _, EXCT_output_dic = res
                # Build the "EXCT" dictionary from the output
                exct_dict = build_exct_dictionary(EXCT_output_dic)
                items[the_index]["EXCT"] = exct_dict

            # Reset for next batch
            async_tasks = []
            async_n = 0

    # --- 5) Write updated JSON back to file ---
    if not output_file_path:
        # If user didn't specify a separate output path, create one
        base_name, ext = os.path.splitext(json_file_path)
        output_file_path = f"{base_name}_extracted_{time.strftime('%Y%m%d_%H%M%S')}.json"

    with open(output_file_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"Extraction completed. Results saved to: {output_file_path}")


######################################
### Example usage (sync entrypoint) ###
######################################
# from S3_EXCT_Pydantic import 

from .S3_EXCT_Pydantic import * # to import all pydantic models and sub-classes
from .S3_EXCT_Pydantic import AIStudy #the main pydantic class
def test_run_extraction_on_json():
    """
    Example of how you might call EXCT_main_JSON in a synchronous context.
    """
    
    # Suppose your JSON is an array of objects, each object has a "text" field:
    json_path = r"C:\Users\LEGION\Documents\GIT\AI-in-Med-Trend\pubmed_data_test\cleaned_pubmed_batch_0_3_302a297f_1500_to_1999.json"
    
    # The user can provide a list of Pydantic classes to use:
    Pydantic_Objects_List = [AIStudy]

    # We run the async method in a sync function for convenience:
    asyncio.run(
        EXCT_main_JSON(
            json_file_path=json_path,
            Pydantic_Objects_List=Pydantic_Objects_List,
            text_key="text",       # which key in each object has the text to extract from
            path_to_list=None,     # if the top-level JSON is already a list
            model_engine="OpenAI_Async",
            model="gpt-4omini",
            max_tokens=2048,
            total_async_n=5,
        )
    )


